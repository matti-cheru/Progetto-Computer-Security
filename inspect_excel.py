import openpyxl

wb = openpyxl.load_workbook('data/csf-pf-to-sp800-53r5-mappings.xlsx')
ws = wb['PF to SP 800-53r5']

relationships = []
for row in range(3, 105):
    subcat = ws.cell(row, 4).value
    if subcat and ":" in str(subcat):
        c = ws.cell(row, 6)
        pat = c.fill.patternType
        theme = c.fill.fgColor.theme if c.fill.fgColor else None
        
        rel = None
        if pat == 'lightGray':
            rel = "Aligned (text adapted)"
        elif pat == 'darkGray':
            rel = "Identical"
        elif pat == 'solid' and theme == 2:
            rel = "Identical"  # Assuming gray
            
        if rel:
            print(f"{subcat.split(':')[0]}: {rel}")
