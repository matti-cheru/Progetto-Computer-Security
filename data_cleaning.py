"""
FASE 1: Data Cleaning e Preparazione
Questo script pulisce i dataset NIST per renderli utilizzabili dal Pandas Agent.

OBIETTIVO: Eliminare intestazioni complesse, celle unite, colonne mal nominate,
e creare dataset puliti e strutturati per query deterministiche.
"""
import pandas as pd
import numpy as np
import os
import warnings
warnings.filterwarnings('ignore')


def clean_csf_to_sp800_53_mapping():
    """
    Pulisce il file di mapping CSF 1.1 -> SP 800-53 Rev 5
    
    Problemi da risolvere:
    - Colonne mal nominate (Unnamed: 1, Unnamed: 2, ecc.)
    - Header complessi su più righe
    - Celle vuote sparse (NaN)
    - Righe di intestazione mescolate con i dati
    """
    print("\n" + "="*70)
    print("🧹 Pulizia: CSF 1.1 → SP 800-53 Rev 5 Mapping")
    print("="*70)
    
    # Leggi il file grezzo
    file_path = "data/csf-pf-to-sp800-53r5-mappings.xlsx"
    df_raw = pd.read_excel(file_path, sheet_name='CSF to SP 800-53r5')
    
    print(f"📥 Dati originali: {df_raw.shape[0]} righe × {df_raw.shape[1]} colonne")
    
    # Analizza i dati per capire la struttura
    # La riga 0 contiene gli header: Function, Category, Subcategory, NIST SP 800-53...
    # Le righe successive contengono i dati gerarchici
    
    # Rinomina le colonne con nomi chiari
    df_raw.columns = ['Function', 'Category', 'Subcategory', 'SP800_53_Controls']
    
    # Rimuovi la prima riga che è un header duplicato
    df = df_raw.iloc[1:].copy()
    
    # Fill forward per le colonne Function e Category (struttura gerarchica)
    # Quando una cella di Function è vuota, significa che appartiene alla Function precedente
    df['Function'] = df['Function'].ffill()
    df['Category'] = df['Category'].ffill()
    
    # Rimuovi righe che non hanno Subcategory (sono solo intestazioni)
    df = df[df['Subcategory'].notna()].copy()
    
    # Pulisci gli spazi bianchi
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].str.strip()
    
    # Estrai l'ID della subcategory (es. "ID.AM-1" da "ID.AM-1: Physical devices...")
    df['Subcategory_ID'] = df['Subcategory'].str.extract(r'^([A-Z]{2}\.[A-Z]{2}-\d+)', expand=False)
    
    # Estrai la descrizione della subcategory (testo dopo ID:)
    df['Subcategory_Description'] = df['Subcategory'].str.replace(
        r'^[A-Z]{2}\.[A-Z]{2}-\d+:\s*', '', regex=True
    )
    
    # Riorganizza le colonne
    df_clean = df[[
        'Function',
        'Category',
        'Subcategory_ID',
        'Subcategory_Description',
        'SP800_53_Controls'
    ]].copy()
    
    # Reset index
    df_clean = df_clean.reset_index(drop=True)
    
    print(f"✅ Dati puliti: {df_clean.shape[0]} righe × {df_clean.shape[1]} colonne")
    print(f"\n📊 Distribuzione:")
    print(f"   - Funzioni uniche: {df_clean['Function'].nunique()}")
    print(f"   - Categorie uniche: {df_clean['Category'].nunique()}")
    print(f"   - Subcategorie uniche: {df_clean['Subcategory_ID'].nunique()}")
    
    # Mostra esempio
    print(f"\n🔍 Esempi di dati puliti:")
    print(df_clean.head(3).to_string())
    
    return df_clean


def clean_pf_to_sp800_53_mapping():
    """
    Pulisce il file di mapping Privacy Framework -> SP 800-53 Rev 5
    """
    print("\n" + "="*70)
    print("🧹 Pulizia: Privacy Framework → SP 800-53 Rev 5 Mapping")
    print("="*70)
    
    file_path = "data/csf-pf-to-sp800-53r5-mappings.xlsx"
    df_raw = pd.read_excel(file_path, sheet_name='PF to SP 800-53r5')
    
    print(f"📥 Dati originali: {df_raw.shape[0]} righe × {df_raw.shape[1]} colonne")
    
    # Usa openpyxl per estrarre il mapping visivo della colonna F prima di pulire
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    ws = wb['PF to SP 800-53r5']
    rel_map = {}
    for r in range(3, ws.max_row + 1):
        subcat_val = ws.cell(row=r, column=4).value
        # Filtriamo le stringhe valide che contengono ':' come 'ID.IM-P1: ...'
        if subcat_val and isinstance(subcat_val, str) and ":" in subcat_val:
            fill = ws.cell(row=r, column=6).fill
            pat = fill.patternType
            theme = fill.fgColor.theme if getattr(fill, 'fgColor', None) else None
            
            # Dal file: lightGray = Aligned, darkGray o solid theme=2 = Identical
            rel_text = pd.NA
            if pat == 'lightGray':
                rel_text = "Aligned - The Privacy Framework Subcategory aligns with the Cybersecurity Framework Subcategory, but the text has been adapted for the Privacy Framework."
            elif pat == 'darkGray' or (pat == 'solid' and str(theme) == '2'):
                rel_text = "Identical - The Privacy Framework Subcategory is identical to the Cybersecurity Framework Subcategory."
                
            rel_map[subcat_val.strip()] = rel_text
            
    # Le prime righe sono header, saltiamole
    df = df_raw.iloc[2:].copy()
    
    # Elimina la prima colonna (che su Excel è solo una barra di colori, vuota)
    df = df.drop(df.columns[0], axis=1)
    
    # Rinomina le colonne basandoti sulla struttura osservata (rimangono 8 colonne)
    df.columns = [
        'Function',
        'Category', 
        'Subcategory',
        'SP800_53_Controls',
        'Relationship_to_CSF',
        'Additional_Info_1',
        'Additional_Info_2',
        'Crosswalk_Link'
    ]
    
    # Fill forward gerarchico
    df['Function'] = df['Function'].ffill()
    df['Category'] = df['Category'].ffill()
    
    # Rimuovi righe senza Subcategory
    df = df[df['Subcategory'].notna()].copy()
    
    # Pulisci spazi
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].str.strip()
            
    # Associa la Relationship estratta tramite openpyxl
    df['Relationship_to_CSF'] = df['Subcategory'].map(rel_map)
    
    # Estrai ID subcategory (es. "ID.IM-P1")
    df['Subcategory_ID'] = df['Subcategory'].str.extract(r'^([A-Z]{2}\.[A-Z]{2,3}-P\d+)', expand=False)
    
    # Estrai descrizione
    df['Subcategory_Description'] = df['Subcategory'].str.replace(
        r'^[A-Z]{2}\.[A-Z]{2,3}-P\d+:\s*', '', regex=True
    )
    
    # Seleziona colonne rilevanti
    df_clean = df[[
        'Function',
        'Category',
        'Subcategory_ID',
        'Subcategory_Description',
        'SP800_53_Controls',
        'Relationship_to_CSF'
    ]].copy()
    
    df_clean = df_clean.reset_index(drop=True)
    
    print(f"✅ Dati puliti: {df_clean.shape[0]} righe × {df_clean.shape[1]} colonne")
    print(f"\n📊 Distribuzione:")
    print(f"   - Funzioni uniche: {df_clean['Function'].nunique()}")
    print(f"   - Categorie uniche: {df_clean['Category'].nunique()}")
    print(f"   - Subcategorie uniche: {df_clean['Subcategory_ID'].nunique()}")
    
    print(f"\n🔍 Esempi di dati puliti:")
    print(df_clean.head(3).to_string())
    
    return df_clean


def clean_sp800_53_catalog():
    """
    Pulisce il catalogo SP 800-53 Rev 5.2.0
    """
    print("\n" + "="*70)
    print("🧹 Pulizia: SP 800-53 Rev 5.2.0 Catalog")
    print("="*70)
    
    file_path = "data/cprt_SP_800_53_5_2_0_03-01-2026.xlsx"
    df_raw = pd.read_excel(file_path, sheet_name='SP 800-53 Rev 5.2.0')
    
    print(f"📥 Dati originali: {df_raw.shape[0]} righe × {df_raw.shape[1]} colonne")
    
    # Rimuovi righe che sono solo reference (senza Control Identifier)
    df = df_raw[df_raw['Control (or Control Enhancement) Identifier'].notna()].copy()
    
    # Rinomina colonne per facilità d'uso
    df.columns = [
        'Control_Family',
        'Control_ID',
        'Control_Name',
        'Control_Statement',
        'Discussion',
        'Related_Controls',
        'Privacy_Baseline',
        'Security_Baseline_Low',
        'Security_Baseline_Moderate',
        'Security_Baseline_High',
        'Reference'
    ]
    
    # Pulisci spazi bianchi
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].str.strip()
    
    # Fill forward per Control_Family
    df['Control_Family'] = df['Control_Family'].ffill()
    
    # Crea colonna che indica se è un control base o enhancement
    df['Is_Enhancement'] = df['Control_ID'].str.contains(r'\(\d+\)', regex=True)
    
    # Estrai il control base (es. "AC-01" da "AC-01(01)")
    df['Base_Control_ID'] = df['Control_ID'].str.extract(r'^([A-Z]{2}-\d+)', expand=False)
    
     # Reset index
    df = df.reset_index(drop=True)
    
    print(f"✅ Dati puliti: {df.shape[0]} righe × {df.shape[1]} colonne")
    print(f"\n📊 Distribuzione:")
    print(f"   - Control Families: {df['Control_Family'].nunique()}")
    print(f"   - Base Controls: {(~df['Is_Enhancement']).sum()}")
    print(f"   - Enhancements: {df['Is_Enhancement'].sum()}")
    
    print(f"\n🔍 Esempi di dati puliti:")
    print(df[['Control_ID', 'Control_Name', 'Control_Family', 'Is_Enhancement']].head(5).to_string())
    
    return df


def main():
    """
    Esegue la pulizia completa di tutti i dataset NIST
    """
    print("="*70)
    print("FASE 1: PULIZIA DATI NIST")
    print("Preparazione per il Pandas Agent - Logical Core")
    print("="*70)
    
    # Crea directory per output
    output_dir = "data/cleaned"
    os.makedirs(output_dir, exist_ok=True)
    
    # Pulisci tutti i dataset
    df_csf_mapping = clean_csf_to_sp800_53_mapping()
    df_pf_mapping = clean_pf_to_sp800_53_mapping()
    df_sp_catalog = clean_sp800_53_catalog()
    
    # Salva in formato CSV (più veloce di Excel per Pandas)
    print("\n" + "="*70)
    print("💾 Salvataggio dati puliti...")
    print("="*70)
    
    csv_files = {
        'csf_to_sp800_53_mapping.csv': df_csf_mapping,
        'pf_to_sp800_53_mapping.csv': df_pf_mapping,
        'sp800_53_catalog.csv': df_sp_catalog
    }
    
    for filename, df in csv_files.items():
        filepath = os.path.join(output_dir, filename)
        df.to_csv(filepath, index=False)
        size_kb = os.path.getsize(filepath) / 1024
        print(f"   ✅ {filename} ({size_kb:.1f} KB)")
    
    # Salva anche in formato Parquet (compresso e ottimizzato) - OPZIONALE
    print("\n💾 Salvataggio formato Parquet (ottimizzato)...")
    
    try:
        for filename, df in csv_files.items():
            parquet_filename = filename.replace('.csv', '.parquet')
            filepath = os.path.join(output_dir, parquet_filename)
            df.to_parquet(filepath, index=False, compression='snappy')
            size_kb = os.path.getsize(filepath) / 1024
            print(f"   ✅ {parquet_filename} ({size_kb:.1f} KB)")
    except ImportError:
        print("   ⚠️  Parquet non disponibile (pyarrow non installato) - usando solo CSV")
    
    print("\n" + "="*70)
    print("✅ PULIZIA COMPLETATA CON SUCCESSO!")
    print("="*70)
    print(f"\n📁 File puliti salvati in: {output_dir}/")
    print("\nI dati sono ora pronti per essere interrogati dal Pandas Agent!")
    print("Nessuna cella unita, nomi colonne chiari, struttura consistente.")


if __name__ == "__main__":
    main()
