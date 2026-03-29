"""
INTERVIEW ENGINE - Phase 1+2: Core Interview Loop + Structured Extraction

Orchestrates the NIST CSF 2.0 compliance interview by:
- Managing compilation sessions (new / resume) in a Compilazioni/ folder
- Pulling the next PENDING subcategory from ProfileManager
- Fetching enriched context from CSF catalog and SP800-53 mappings
- Building a professional, contextualized question via LLM
- Extracting structured data from the user's free-text answer
- Saving progress incrementally to the session's profile CSV

Supports verbose and non-verbose modes with full JSON logging.
"""
import os
import re
import json
import shutil
from typing import Dict, List, Optional, Any, Set
from datetime import datetime

import pandas as pd
from openai import OpenAI
from dotenv import load_dotenv

load_dotenv()

from profile_manager import ProfileManager
from pandas_agent_manual import ManualPandasAgent


# ═══════════════════════════════════════════════════════════════════
#  DATA PATHS
# ═══════════════════════════════════════════════════════════════════
DATA_DIR = os.path.join(os.path.dirname(__file__), "data", "cleaned")
CSF_CATALOG_PATH = os.path.join(DATA_DIR, "csf_2_0_catalog.csv")
CSF_SP800_MAPPING_PATH = os.path.join(DATA_DIR, "csf_to_sp800_53_mapping.csv")
SP800_CATALOG_PATH = os.path.join(DATA_DIR, "sp800_53_catalog.csv")


# ═══════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════

COMPILAZIONI_DIR = os.path.join(os.getcwd(), "Compilazioni")
PROFILE_FILENAME_PREFIX = "profile_"
PROFILE_FILENAME_SUFFIX = ".csv"


class InterviewEngine:
    """
    Core engine for the NIST CSF 2.0 compliance interview.

    Connects ProfileManager (state) with the LLM to run a
    conversational loop that progressively fills the organizational profile.
    """

    # ── Subcategory scope limiter (for testing) ──────────────────────
    # Set to None to process ALL subcategories.
    SCOPE_SUBCATEGORIES: Optional[Set[str]] = None

    # ── Profile columns that the LLM must fill ──────────────────────
    CURRENT_PROFILE_COLUMNS = [
        "Included_in_Profile",
        "Rationale",
        "Current_Priority",
        "Current_Status",
        "Current_Policies_Processes_Procedures",
        "Current_Internal_Practices",
        "Current_Roles_and_Responsibilities",
        "Current_Selected_Informative_References",
        "Current_Artifacts_and_Evidence",
    ]

    TARGET_PROFILE_COLUMNS = [
        "Target_Priority",
        "Target_CSF_Tier",
        "Target_Policies_Processes_Procedures",
        "Target_Internal_Practices",
        "Target_Roles_and_Responsibilities",
        "Target_Selected_Informative_References",
    ]

    ALL_PROFILE_COLUMNS = CURRENT_PROFILE_COLUMNS + TARGET_PROFILE_COLUMNS + ["Notes", "Considerations"]

    # ── LLM parameters ──────────────────────────────────────────────
    # The extraction call needs very high max_tokens because reasoning
    # models (like gpt-oss) can consume thousands of tokens on internal
    # reasoning loops before producing the visible JSON output.
    EXTRACTION_MAX_TOKENS = 4096
    EXTRACTION_RETRIES = 3

    def __init__(
        self,
        base_url: str = "https://gpustack.ing.unibs.it/v1",
        interview_model_name: str = "qwen3",
        pandas_agent_model_name: str = "gpt-oss",
        api_key: Optional[str] = None,
        verbose: bool = True,
        log_dir: Optional[str] = None,
    ):
        """
        Args:
            base_url:   LLM endpoint URL
            interview_model_name: Fast conversational model (e.g. qwen3, phi4)
            pandas_agent_model_name: Heavy reasoning model (e.g. gpt-oss)
            api_key:    API key (falls back to env GPUSTACK_API_KEY)
            verbose:    If True, print full internal traces to console
            log_dir:    Directory for JSON logs (auto-created per run)
        """
        if api_key is None:
            
            api_key = os.environ.get("GPUSTACK_API_KEY")
            
            if not api_key:
                raise ValueError("Nessuna API Key trovata! Assicurati di impostare GPUSTACK_API_KEY nel file .env")

        self.verbose = verbose
        # The main model used for questions and structured extraction
        self.model_name = interview_model_name
        self.pandas_agent_model_name = pandas_agent_model_name
        self.client = OpenAI(base_url=base_url, api_key=api_key)

        # ProfileManager will be initialized during session selection
        self.manager: Optional[ProfileManager] = None

        # Run-level logging directory (set during session selection)
        self.log_dir = log_dir

        # Accumulated interview log (saved at each turn)
        self._turn_logs: List[Dict[str, Any]] = []

        # ── Load reference data for catalog context ──────────────
        self._csf_catalog = self._load_csv_safe(CSF_CATALOG_PATH)
        self._sp800_mapping = self._load_csv_safe(CSF_SP800_MAPPING_PATH)
        self._sp800_catalog = self._load_csv_safe(SP800_CATALOG_PATH)

        # ── Pandas Agent for catalog context queries ───────────
        self._pandas_agent = ManualPandasAgent(
            base_url=base_url,
            model_name=self.pandas_agent_model_name,
            api_key=api_key,
            temperature=0.0,
            verbose=verbose,   # follow the engine's verbosity
        )

    @staticmethod
    def _load_csv_safe(path: str) -> Optional[pd.DataFrame]:
        """Load a CSV file, returning None if it doesn't exist."""
        if os.path.exists(path):
            return pd.read_csv(path, dtype=str)
        return None

    # ═══════════════════════════════════════════════════════════════
    #  PUBLIC API
    # ═══════════════════════════════════════════════════════════════

    def start(self):
        """
        Main entry point.
        1. Ask user: new compilation or resume existing
        2. Run the interview loop
        """
        self._print_always(
            "\n" + "=" * 62
            + "\n🛡️  NIST CSF 2.0 — Interview Engine"
            + f"\n   Model: {self.model_name}"
            + "\n" + "=" * 62
        )

        # ── Session selection ────────────────────────────────────
        session_path = self._session_selection()
        if session_path is None:
            self._print_always("\n👋 Goodbye!")
            return

        # Initialize ProfileManager on the selected session file
        # verbose=False suppresses internal "Stato salvato" messages
        self.manager = ProfileManager(
            profile_name=os.path.basename(session_path),
            save_dir=os.path.dirname(session_path),
            verbose=False,
        )

        # Setup log directory alongside the profile
        if self.log_dir is None:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.log_dir = os.path.join(
                os.path.dirname(session_path), f"logs_{ts}"
            )
        os.makedirs(self.log_dir, exist_ok=True)

        self._print_always(f"   Profile: {session_path}")
        self._print_always(f"   Logs:    {self.log_dir}")
        self._print_always(
            f"   Scope:   {sorted(self.SCOPE_SUBCATEGORIES) if self.SCOPE_SUBCATEGORIES else 'ALL subcategories'}"
        )

        # ── Interview loop or Revision ─────────────────────────────
        progress = self.manager.get_progress_summary()
        if progress["completed"] > 0:
            self._print_always("\n" + "="*60)
            self._print_always("🔄 Resume Options")
            self._print_always("="*60)
            self._print_always("  [1] Resume normal progression (from where you left off)")
            self._print_always("  [2] Revise an existing completed subcategory")
            
            while True:
                resume_choice = input("\nSelect option [1-2] (default: 1) ▶ ").strip()
                if not resume_choice or resume_choice == "1":
                    self._run_interview_loop()
                    break
                elif resume_choice == "2":
                    self._revise_subcategory_loop()
                    # After finishing revision, ask if they want to resume normal progression
                    cont = input("\nDo you want to continue with normal interview progression? (y/n) ▶ ").strip().lower()
                    if cont in ["y", "yes"]:
                        self._run_interview_loop()
                    break
                else:
                    self._print_always("⚠️ Invalid choice. Select 1 or 2.")
        else:
            self._run_interview_loop()

    # ═══════════════════════════════════════════════════════════════
    #  SESSION MANAGEMENT
    # ═══════════════════════════════════════════════════════════════

    def _session_selection(self) -> Optional[str]:
        """
        Prompts user to start a new compilation or resume an existing one.
        Returns the path to the profile CSV, or None to exit.
        """
        os.makedirs(COMPILAZIONI_DIR, exist_ok=True)

        # Find existing compilations
        existing = self._list_compilations()

        self._print_always(
            "\n╔════════════════════════════════════════════════════════════╗"
            "\n║           Select Compilation Mode                          ║"
            "\n╠════════════════════════════════════════════════════════════╣"
        )

        if existing:
            self._print_always(
                "║                                                            ║"
                "\n║  [N] Start a NEW compilation                               ║"
                "\n║  [R] Resume an existing compilation                        ║"
                "\n║  [E] Export an existing compilation to Excel               ║"
                "\n║  [Q] Quit                                                  ║"
                "\n║                                                            ║"
                "\n╚════════════════════════════════════════════════════════════╝"
            )
        else:
            self._print_always(
                "║                                                           ║"
                "\n║  No existing compilations found.                           ║"
                "\n║  A new compilation will be created.                        ║"
                "\n║                                                            ║"
                "\n╚════════════════════════════════════════════════════════════╝"
            )
            return self._create_new_session()

        while True:
            choice = input("\nYour choice [N/R/E/Q] ▶ ").strip().upper()

            if choice == "Q":
                return None

            if choice == "N":
                return self._create_new_session()

            if choice == "R":
                return self._resume_session(existing)
                
            if choice == "E":
                self._export_session_to_excel(existing)
                return self._session_selection() # Reload menu after export

            self._print_always("⚠️  Invalid choice. Please enter N, R, E, or Q.")

    def _list_compilations(self) -> List[Dict[str, Any]]:
        """Lists all existing compilation files in Compilazioni/."""
        compilations = []
        if not os.path.exists(COMPILAZIONI_DIR):
            return compilations

        for fname in sorted(os.listdir(COMPILAZIONI_DIR)):
            if fname.startswith(PROFILE_FILENAME_PREFIX) and fname.endswith(PROFILE_FILENAME_SUFFIX):
                fpath = os.path.join(COMPILAZIONI_DIR, fname)
                try:
                    # Quick peek at progress
                    df = pd.read_csv(fpath, usecols=["Completion_Status"], dtype={"Completion_Status": str})
                    total = len(df)
                    done = (df["Completion_Status"] == "DONE").sum()
                    pct = (done / total * 100) if total > 0 else 0

                    # Extract timestamp from filename
                    ts_match = re.search(r"(\d{8}_\d{6})", fname)
                    ts_str = ts_match.group(1) if ts_match else "unknown"

                    compilations.append({
                        "filename": fname,
                        "path": fpath,
                        "timestamp": ts_str,
                        "done": done,
                        "total": total,
                        "percentage": pct,
                    })
                except Exception:
                    pass  # Skip corrupt files

        return compilations

    def _create_new_session(self) -> str:
        """Creates a fresh profile CSV in Compilazioni/ with a timestamp."""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"{PROFILE_FILENAME_PREFIX}{ts}{PROFILE_FILENAME_SUFFIX}"
        dest_path = os.path.join(COMPILAZIONI_DIR, fname)

        # Use ProfileManager to create a pristine copy from the catalog
        temp_manager = ProfileManager(verbose=False)
        temp_manager.create_fresh_copy(dest_path)

        self._print_always(f"\n✅ New compilation created: {fname}")
        return dest_path

    def _resume_session(self, compilations: List[Dict[str, Any]]) -> Optional[str]:
        """Shows existing compilations and lets user pick one."""
        self._print_always("\n📂 Existing compilations:\n")
        for i, comp in enumerate(compilations, 1):
            dt = comp["timestamp"]
            # Format timestamp nicely
            try:
                dt_formatted = datetime.strptime(dt, "%Y%m%d_%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                dt_formatted = dt
            self._print_always(
                f"   [{i}] {comp['filename']}"
                f"\n       Created: {dt_formatted}"
                f"\n       Progress: {comp['done']}/{comp['total']} ({comp['percentage']:.0f}%)"
                f"\n"
            )

        while True:
            choice = input(f"Select compilation [1-{len(compilations)}] or Q to go back ▶ ").strip()

            if choice.upper() == "Q":
                return self._session_selection()

            try:
                idx = int(choice) - 1
                if 0 <= idx < len(compilations):
                    selected = compilations[idx]
                    self._print_always(f"\n✅ Resuming: {selected['filename']}")
                    return selected["path"]
                else:
                    self._print_always(f"⚠️  Please enter a number between 1 and {len(compilations)}.")
            except ValueError:
                self._print_always("⚠️  Invalid input. Enter a number or Q.")

    def _export_session_to_excel(self, compilations: List[Dict[str, Any]]):
        """Lets the user select a compilation and exports it to the official NIST Excel template."""
        self._print_always("\n📂 Select a compilation to export:\n")
        for i, comp in enumerate(compilations, 1):
            dt = comp["timestamp"]
            try:
                dt_formatted = datetime.strptime(dt, "%Y%m%d_%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                dt_formatted = dt
            self._print_always(
                f"   [{i}] {comp['filename']}"
                f"\n       Created: {dt_formatted}"
                f"\n       Progress: {comp['done']}/{comp['total']} ({comp['percentage']:.0f}%)"
                f"\n"
            )

        while True:
            choice = input(f"Select compilation [1-{len(compilations)}] or Q to cancel ▶ ").strip()

            if choice.upper() == "Q":
                return

            try:
                idx = int(choice) - 1
                if 0 <= idx < len(compilations):
                    src_csv = compilations[idx]["path"]
                    dt_stamp = compilations[idx]["timestamp"]
                    out_excel = os.path.join(COMPILAZIONI_DIR, f"profile_exported_{dt_stamp}.xlsx")
                    self._perform_excel_export(src_csv, out_excel)
                    return
            except ValueError:
                pass
            self._print_always("⚠️  Invalid choice.")

    def _perform_excel_export(self, csv_path: str, out_excel_path: str):
        """
        Reads the filled CSV and maps its columns directly over the template.
        Uses openpyxl to write without destroying the file formatting.
        """
        template_path = "data/CSF 2.0 Organizational Profile Template.xlsx"
        if not os.path.exists(template_path):
            self._print_always(f"❌ Error: The template file '{template_path}' does not exist.")
            return
            
        try:
            import openpyxl
        except ImportError:
            self._print_always(f"❌ Error: 'openpyxl' module is required for Excel export. Please install it with 'pip install openpyxl'.")
            return
            
        self._print_always("\n⏳ Exporting profile to NIST Excel Template...")
        
        # Load the compilations CSV
        df_csv = pd.read_csv(csv_path).fillna("")
        
        # Load the workbook
        wb = openpyxl.load_workbook(template_path)
        if "Current and Target Profile" not in wb.sheetnames:
            self._print_always("❌ Error: 'Current and Target Profile' sheet not found in the template.")
            return
            
        sheet = wb["Current and Target Profile"]
        
        # The template has headers on row 1
        headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
        
        col_mapping = {
            "Subcategory_ID": "CSF Outcome (Function, Category, or Subcategory)",
            "Subcategory_Description": "CSF Outcome Description",
            "Included_in_Profile": "Included in Profile?",
            "Rationale": "Rationale",
            
            "Current_Priority": "Current Priority",
            "Current_Status": "Current Status",
            "Current_Policies_Processes_Procedures": "Current Policies, Processes, and Procedures",
            "Current_Internal_Practices": "Current Internal Practices",
            "Current_Roles_and_Responsibilities": "Current Roles and Responsibilities",
            "Current_Selected_Informative_References": "Current Selected Informative References",
            "Current_Artifacts_and_Evidence": "Current Artifacts and Evidence",
            
            "Target_Priority": "Target Priority",
            "Target_CSF_Tier": "Target CSF Tier",
            "Target_Policies_Processes_Procedures": "Target Policies, Processes, and Procedures",
            "Target_Internal_Practices": "Target Internal Practices",
            "Target_Roles_and_Responsibilities": "Target Roles and Responsibilities",
            "Target_Selected_Informative_References": "Target Selected Informative References"
        }
        
        outcome_col_idx = headers.get("CSF Outcome (Function, Category, or Subcategory)")
        if not outcome_col_idx:
            self._print_always("❌ Error: Could not find 'CSF Outcome' column in template.")
            return
            
        id_to_row = {}
        for r_idx in range(2, sheet.max_row + 1):
            val = sheet.cell(row=r_idx, column=outcome_col_idx).value
            if val:
                val_str = str(val).strip()
                id_to_row[val_str] = r_idx
                norm_val = self._normalize_subcategory_id(val_str)
                if norm_val:
                    id_to_row[norm_val] = r_idx

        rows_written = 0
        for _, csv_row in df_csv.iterrows():
            sid = str(csv_row.get("Subcategory_ID", "")).strip()
            norm_sid = self._normalize_subcategory_id(sid)
            
            target_row = id_to_row.get(sid) or id_to_row.get(norm_sid)
            
            if not target_row:
                continue
                
            for csv_col, excel_header in col_mapping.items():
                if csv_col in ["Subcategory_ID", "Subcategory_Description"]:
                    continue
                    
                target_col_idx = headers.get(excel_header)
                if target_col_idx:
                    val = csv_row.get(csv_col, "")
                    if val and val != "Not specified":
                        sheet.cell(row=target_row, column=target_col_idx).value = val
            rows_written += 1
            
        wb.save(out_excel_path)
        self._print_always(f"✅ Export completed! Written {rows_written} subcategories.")
        self._print_always(f"📂 Saved to: {out_excel_path}\n")

    # ═══════════════════════════════════════════════════════════════
    #  INTERVIEW LOOP
    # ═══════════════════════════════════════════════════════════════

    def _run_interview_loop(self):
        """
        Runs the interview loop until all in-scope subcategories are
        DONE, the user types /quit, or there are no more PENDING items.
        """
        self._print_always(
            "\n╔════════════════════════════════════════════════════════════╗"
            "\n║  Interview Started                                         ║"
            "\n╠════════════════════════════════════════════════════════════╣"
            "\n║  Commands:                                                 ║"
            "\n║   /progress  — Show completion progress                    ║"
            "\n║   /skip      — Skip the current subcategory                ║"
            "\n║   /quit      — Save and exit (resume later)                ║"
            "\n╚════════════════════════════════════════════════════════════╝"
        )

        self._show_progress()

        while True:
            row = self._get_next_in_scope()
            if row is None:
                self._print_always(
                    "\n✅ All in-scope subcategories have been completed!"
                )
                self._show_progress()
                break

            sid = row["Subcategory_ID"]
            self._print_always(f"\n{'─' * 60}")
            self._print_always(f"📋 Subcategory: {sid}")
            self._print_always(f"{'─' * 60}")
            # Fetch enriched catalog context for this subcategory
            catalog_ctx, catalog_ctx_trace = self._fetch_catalog_context(sid)

            # =========================================================
            # PHASE 1: CURRENT STATE
            # =========================================================
            self._print_always(f"\n[PHASE 1 - CURRENT STATE]")
            
            # --- 1) Build CURRENT question via LLM ---
            question, question_trace = self._build_question(row, catalog_ctx)

            # --- 2) Ask user ---
            self._print_always(f"\n{question}\n")
            user_input = input("Your answer (or /skip, /quit, /progress) ▶ ").strip()

            # Handle commands
            if user_input.lower() == "/quit":
                # Revert to PENDING since we haven't processed it
                self.manager.update_row(
                    sid, {"Completion_Status": "PENDING"}
                )
                self._print_always("\n💾 Progress saved. You can resume later.")
                self._show_progress()
                break

            if user_input.lower() in ["/skip", "skip"]:
                self.manager.update_row(
                    sid, {
                        "Completion_Status": "DONE",
                        "Included_in_Profile": "Skipped",
                        "Notes": "User skipped this subcategory.",
                    }
                )
                self._print_always(f"⏭️  Skipped {sid}")
                continue

            if user_input.lower() == "/progress":
                self._show_progress()
                # Revert to PENDING so it gets picked up again
                self.manager.update_row(
                    sid, {"Completion_Status": "PENDING"}
                )
                continue

            if not user_input:
                self._print_always("⚠️  Empty answer. Please provide a response or use /skip.")
                self.manager.update_row(
                    sid, {"Completion_Status": "PENDING"}
                )
                continue

            # --- 3) Extract CURRENT response via LLM ---
            current_extracted, current_extraction_trace = self._extract_response(
                row, user_input, catalog_ctx
            )
            
            while True:
                # Show extracted data to user for transparency
                self._print_always(f"\n📊 Extracted CURRENT profile data for {sid}:")
                for col, val in current_extracted.items():
                    if col != "Completion_Status" and val:
                        self._print_always(f"   • {col}: {val}")
                        
                self._print_always("\nAre these details correct? (Press Enter to confirm, or type your corrections below)")
                feedback = input("Confirm or refine ▶ ").strip()
                if not feedback or feedback.lower() in ["yes", "y", "ok", "confirm"]:
                    break
                
                self._print_always("\n🔄 Refining extraction based on your feedback...")
                current_extracted, new_trace = self._extract_response(
                    row, user_input, catalog_ctx, previous_extracted=current_extracted, feedback=feedback
                )
                if "attempts" in current_extraction_trace and "attempts" in new_trace:
                    current_extraction_trace["attempts"].extend(new_trace["attempts"])
                    
            # --- 4) Save partial progress ---
            current_extracted["Completion_Status"] = "IN_PROGRESS"
            self.manager.update_row(sid, current_extracted)

            # =========================================================
            # PHASE 2: TARGET STATE
            # =========================================================
            self._print_always(
                f"\n{'='*70}\n[PHASE 2 - TARGET STATE] Subcategory: {sid}\n{'='*70}"
            )
            
            # --- 5) Build TARGET question via LLM ---
            target_question, target_question_trace = self._build_target_question(
                row, catalog_ctx, current_extracted
            )
            
            # --- 6) Ask user ---
            self._print_always(f"\n{target_question}\n")
            target_input = input("Your answer (or /skip) ▶ ").strip()
            
            if target_input.lower() in ["/quit", "quit", "exit"]:
                self._print_always("\nStopping interview. Progress has been saved in IN_PROGRESS state.")
                break
                
            if target_input.lower() in ["/skip", "skip"]:
                self.manager.update_row(
                    sid, {
                        "Completion_Status": "DONE",
                        "Included_in_Profile": current_extracted.get("Included_in_Profile", "Yes"),
                        "Notes": "User skipped TARGET phase.",
                    }
                )
                self._print_always(f"⏭️  Skipped TARGET phase for {sid}")
                continue
            
            if not target_input:
                self._print_always("Empty answer for Target. Will proceed with empty info...\n")
            
            # --- 7) Extract TARGET response via LLM ---
            target_extracted, target_extraction_trace = self._extract_target_response(
                row, target_input, catalog_ctx
            )
            
            while True:
                self._print_always(f"\n📊 Extracted TARGET profile data for {sid}:")
                for col, val in target_extracted.items():
                    if col != "Completion_Status" and val:
                        self._print_always(f"   • {col}: {val}")
                        
                self._print_always("\nAre these details correct? (Press Enter to confirm, or type your corrections below)")
                feedback = input("Confirm or refine ▶ ").strip()
                if not feedback or feedback.lower() in ["yes", "y", "ok", "confirm"]:
                    break
                
                self._print_always("\n🔄 Refining extraction based on your feedback...")
                target_extracted, new_trace = self._extract_target_response(
                    row, target_input, catalog_ctx, previous_extracted=target_extracted, feedback=feedback
                )
                if "attempts" in target_extraction_trace and "attempts" in new_trace:
                    target_extraction_trace["attempts"].extend(new_trace["attempts"])
            
            # --- 8) Finish and Save ---
            target_extracted["Completion_Status"] = "DONE"
            self.manager.update_row(sid, target_extracted)
            self._print_always(f"\n✅ {sid} saved successfully.")

            # --- 9) Log the full dual-phase interaction ---
            turn_log = {
                "turn": len(self._turn_logs) + 1,
                "timestamp": datetime.now().isoformat(),
                "subcategory_id": sid,
                "row_context": {
                    "Function": row.get("Function", ""),
                    "Category": row.get("Category", ""),
                    "Subcategory_Description": row.get("Subcategory_Description", ""),
                    "Implementation_Examples": row.get("Implementation_Examples", ""),
                },
                "catalog_context": catalog_ctx_trace if catalog_ctx else None,
                "current_phase": {
                    "question_generation": question_trace,
                    "user_answer": user_input,
                    "extraction": current_extraction_trace,
                    "extracted_data": current_extracted,
                },
                "target_phase": {
                    "question_generation": target_question_trace,
                    "user_answer": target_input,
                    "extraction": target_extraction_trace,
                    "extracted_data": target_extracted,
                }
            }
            self._turn_logs.append(turn_log)
            self._save_run_log()

    def _revise_subcategory_loop(self):
        """
        Allows the user to enter a continuous loop to revise already completed subcategories.
        Fetches the context, loads the current CSV row, and runs the confirmation loops.
        """
        while True:
            self._print_always("\n" + "="*60)
            sid = input("Enter Subcategory ID to revise (e.g. ID.AM-01) or input '/quit' to exit revision mode ▶ ").strip()
            
            if sid.lower() in ["/quit", "quit", "q", "exit"]:
                self._print_always("Exiting revision mode.")
                break
            
            if not sid:
                continue
                
            # Allow common typos like id.am-1 or id.am-01 by normalizing
            df = self.manager.df
            row_df = df[df['Subcategory_ID'].str.lower() == sid.lower()]
            if row_df.empty:
                self._print_always(f"❌ Subcategory '{sid}' not found inside the current profile's scope.")
                continue
                
            # Get canonical casing
            sid = row_df.iloc[0]['Subcategory_ID']
            row = row_df.iloc[0].to_dict()
            
            if row.get("Completion_Status") == "PENDING":
                self._print_always(f"⚠️  La Sottocategoria '{sid}' non è ancora stata compilata! Ritorna al flusso normale per rispondere alle domande.")
                continue
            
            self._print_always(f"\n{'─' * 60}")
            self._print_always(f"🛠️ REVISING Subcategory: {sid}")
            self._print_always(f"{'─' * 60}")
            
            # Fetch context via Pandas Agent
            catalog_ctx, catalog_ctx_trace = self._fetch_catalog_context(sid)
            dummy_answer = "User is revising previously saved data."
            
            # =========================================================
            # PHASE 1: CURRENT STATE REVISION
            # =========================================================
            self._print_always(f"\n[PHASE 1 - CURRENT STATE REVISION]")
            
            current_extracted = {}
            for col in self.CURRENT_PROFILE_COLUMNS:
                current_extracted[col] = row.get(col, "Not specified")
                
            while True:
                self._print_always(f"\n📊 Saved CURRENT profile data for {sid}:")
                for col, val in current_extracted.items():
                    if col != "Completion_Status" and val:
                        self._print_always(f"   • {col}: {val}")
                        
                self._print_always("\nAre these details correct? (Press Enter to confirm, or type your corrections below)")
                feedback = input("Confirm or refine ▶ ").strip()
                if not feedback or feedback.lower() in ["yes", "y", "ok", "confirm"]:
                    break
                
                self._print_always("\n🔄 Refining CURRENT extraction based on your feedback...")
                current_extracted, new_trace = self._extract_response(
                    row, dummy_answer, catalog_ctx, previous_extracted=current_extracted, feedback=feedback
                )
                
            current_extracted["Completion_Status"] = "IN_PROGRESS"
            self.manager.update_row(sid, current_extracted)
            
            # =========================================================
            # PHASE 2: TARGET STATE REVISION
            # =========================================================
            self._print_always(f"\n[PHASE 2 - TARGET STATE REVISION]")
            
            # Re-fetch the row to get combined updates
            row = self.manager.df[self.manager.df['Subcategory_ID'] == sid].iloc[0].to_dict()
            target_extracted = {}
            for col in self.TARGET_PROFILE_COLUMNS:
                target_extracted[col] = row.get(col, "Not specified")
                
            while True:
                self._print_always(f"\n📊 Saved TARGET profile data for {sid}:")
                for col, val in target_extracted.items():
                    if col != "Completion_Status" and val:
                        self._print_always(f"   • {col}: {val}")
                        
                self._print_always("\nAre these details correct? (Press Enter to confirm, or type your corrections below)")
                feedback = input("Confirm or refine ▶ ").strip()
                if not feedback or feedback.lower() in ["yes", "y", "ok", "confirm"]:
                    break
                
                self._print_always("\n🔄 Refining TARGET extraction based on your feedback...")
                target_extracted, new_trace = self._extract_target_response(
                    row, dummy_answer, catalog_ctx, previous_extracted=target_extracted, feedback=feedback
                )
                
            target_extracted["Completion_Status"] = "DONE"
            self.manager.update_row(sid, target_extracted)
            self._print_always(f"\n✅ Revision for {sid} completed and saved!")

    # ═══════════════════════════════════════════════════════════════
    #  INTERNAL — Scope filtering
    # ═══════════════════════════════════════════════════════════════

    def _get_next_in_scope(self) -> Optional[Dict]:
        """
        Returns the next PENDING row that falls within SCOPE_SUBCATEGORIES.
        If SCOPE_SUBCATEGORIES is None, returns any next PENDING row.
        """
        if self.SCOPE_SUBCATEGORIES is None:
            return self.manager.get_next_pending()

        pending = self.manager.df[
            self.manager.df["Completion_Status"] == "PENDING"
        ]
        in_scope = pending[
            pending["Subcategory_ID"].isin(self.SCOPE_SUBCATEGORIES)
        ]
        if in_scope.empty:
            return None
        return in_scope.iloc[0].to_dict()

    # ═══════════════════════════════════════════════════════════════
    #  INTERNAL — Catalog context enrichment
    # ═══════════════════════════════════════════════════════════════

    @staticmethod
    def _normalize_subcategory_id(sub_id: str) -> str:
        """
        Normalize subcategory ID between CSF 2.0 format (ID.AM-01)
        and the mapping file format (ID.AM-1), stripping leading zeros.
        Returns the version WITHOUT leading zeros (e.g. ID.AM-1).
        """
        # "ID.AM-01" → "ID.AM-1"
        return re.sub(r'-0*(\d+)$', r'-\1', sub_id)

    @staticmethod
    def _normalize_control_id(ctrl_id: str) -> str:
        """
        Normalize SP800-53 control IDs.
        Mapping uses 'CM-8', catalog uses 'CM-08'.
        Returns the version WITH leading zeros (e.g. CM-08).
        """
        # "CM-8" → "CM-08"  (pad to 2 digits)
        def _pad(m):
            return f"-{m.group(1).zfill(2)}"
        return re.sub(r'-(?!\d{2})(?!\()(\d+)', _pad, ctrl_id.strip())

    def _fetch_catalog_context(
        self, subcategory_id: str
    ) -> tuple:
        """
        Fetches enriched context for a subcategory using the Pandas Agent.
        Each lookup goes through the full ReAct pipeline so the user can
        see: prompt → reasoning → generated code → execution result.

        Steps:
        1. CSF catalog: description + examples from csf_2_0_catalog.csv
        2. SP800-53 mapping: control IDs from csf_to_sp800_53_mapping.csv
        3. SP800-53 details: control names + statements from sp800_53_catalog.csv

        Returns:
            (context_dict, trace_dict)
        """
        ctx: Dict[str, Any] = {
            "catalog_description": "",
            "catalog_examples": "",
            "sp800_controls": [],
        }
        trace: Dict[str, Any] = {
            "subcategory_id": subcategory_id,
            "pandas_agent_queries": [],
        }

        # ---- Step 1: CSF Catalog lookup via Pandas Agent ----
        if self._csf_catalog is not None:
            question_1 = (
                f"Get the Subcategory_Description and Implementation_Examples "
                f"for the row where Subcategory_ID == '{subcategory_id}'. "
                f"Return them as a string."
            )
            self._print_verbose(f"\n📚 [Pandas Agent] Step 1: CSF Catalog lookup for {subcategory_id}")

            agent_result_1 = self._pandas_agent.query(
                df=self._csf_catalog,
                question=question_1,
                max_iterations=3,
            )

            trace["pandas_agent_queries"].append({
                "step": "csf_catalog_lookup",
                "question": question_1,
                "success": agent_result_1["success"],
                "answer": agent_result_1.get("answer", ""),
                "iterations": agent_result_1["iterations"],
                "history": agent_result_1.get("history", []),
            })


            cat_rows = self._csf_catalog[
                self._csf_catalog["Subcategory_ID"] == subcategory_id
            ]
            if not cat_rows.empty:
                row = cat_rows.iloc[0]
                desc = row.get("Subcategory_Description", "")
                examples_raw = row.get("Implementation_Examples", "")
                if pd.isna(desc):
                    desc = ""
                if pd.isna(examples_raw):
                    examples_raw = ""
                ctx["catalog_description"] = desc
                ctx["catalog_examples"] = examples_raw

        # ---- Step 2: SP800-53 mapping lookup via Pandas Agent ----
        normalized_id = self._normalize_subcategory_id(subcategory_id)
        raw_controls = []

        if self._sp800_mapping is not None:
            question_2 = (
                f"Get the SP800_53_Controls column value for the row where "
                f"Subcategory_ID == '{normalized_id}'. Return just the value."
            )
            self._print_verbose(f"\n📚 [Pandas Agent] Step 2: SP800-53 mapping for {normalized_id}")

            agent_result_2 = self._pandas_agent.query(
                df=self._sp800_mapping,
                question=question_2,
                max_iterations=3,
            )

            trace["pandas_agent_queries"].append({
                "step": "sp800_mapping_lookup",
                "normalized_id": normalized_id,
                "question": question_2,
                "success": agent_result_2["success"],
                "answer": agent_result_2.get("answer", ""),
                "iterations": agent_result_2["iterations"],
                "history": agent_result_2.get("history", []),
            })

            # Direct lookup for reliable data extraction
            map_rows = self._sp800_mapping[
                self._sp800_mapping["Subcategory_ID"] == normalized_id
            ]
            if not map_rows.empty:
                controls_str = map_rows.iloc[0].get("SP800_53_Controls", "")
                if pd.isna(controls_str):
                    controls_str = ""
                raw_controls = [c.strip() for c in controls_str.split(",") if c.strip()]

        # ---- Step 3: SP800-53 catalog detail lookup via Pandas Agent ----
        if self._sp800_catalog is not None and raw_controls:
            # Normalize control IDs for lookup
            norm_controls = [self._normalize_control_id(c) for c in raw_controls]
            controls_filter = ", ".join(f"'{c}'" for c in norm_controls)

            question_3 = (
                f"Get Control_ID, Control_Name, and Control_Statement for rows where "
                f"Control_ID is in [{controls_filter}]. Return as a text list of dictionaries (using .to_dict('records') or similar format)."
            )
            self._print_verbose(f"\n📚 [Pandas Agent] Step 3: SP800-53 catalog details for {norm_controls}")

            agent_result_3 = self._pandas_agent.query(
                df=self._sp800_catalog,
                question=question_3,
                max_iterations=3,
            )

            trace["pandas_agent_queries"].append({
                "step": "sp800_catalog_lookup",
                "control_ids_queried": norm_controls,
                "question": question_3,
                "success": agent_result_3["success"],
                "answer": agent_result_3.get("answer", ""),
                "iterations": agent_result_3["iterations"],
                "history": agent_result_3.get("history", []),
            })

            # Direct lookup for reliable data
            control_details = []
            for norm_ctrl in norm_controls:
                ctrl_rows = self._sp800_catalog[
                    self._sp800_catalog["Control_ID"] == norm_ctrl
                ]
                if not ctrl_rows.empty:
                    crow = ctrl_rows.iloc[0]
                    detail = {
                        "control_id": norm_ctrl,
                        "control_name": crow.get("Control_Name", ""),
                        "control_statement": crow.get("Control_Statement", "")[:500],
                    }
                    control_details.append(detail)
            ctx["sp800_controls"] = control_details

        self._print_verbose(f"\n📚 [Catalog Context Summary] {subcategory_id}:")
        self._print_verbose(f"   Description: {ctx['catalog_description'][:100]}...")
        self._print_verbose(f"   Examples: {'Yes' if ctx['catalog_examples'] else 'No'}")
        self._print_verbose(
            f"   SP800-53 Controls: {[c['control_id'] for c in ctx['sp800_controls']]}"
        )

        return ctx, trace

    # ═══════════════════════════════════════════════════════════════
    #  INTERNAL — Question building
    # ═══════════════════════════════════════════════════════════════

    def _build_question(
        self, row: Dict, catalog_ctx: Optional[Dict] = None
    ) -> tuple:
        """
        Uses the LLM to formulate a clear, professional interview question
        from the subcategory's catalog data, enriched with SP800-53 context.

        Returns:
            (question_text, trace_dict)
        """
        subcategory_id = row.get("Subcategory_ID", "")
        category = row.get("Category", "")
        description = row.get("Subcategory_Description", "")
        examples = row.get("Implementation_Examples", "")

        # Handle NaN in examples
        if pd.isna(examples) or str(examples).strip().lower() == "nan":
            examples = "No specific implementation examples available."

        # Build SP800-53 controls context
        sp800_section = ""
        if catalog_ctx and catalog_ctx.get("sp800_controls"):
            controls_text = "\n".join(
                f"  - {c['control_id']} ({c['control_name']}): {c['control_statement'][:200]}"
                for c in catalog_ctx["sp800_controls"]
            )
            sp800_section = f"\nRELATED NIST SP 800-53 CONTROLS:\n{controls_text}\n"

        prompt = f"""You are a professional cybersecurity auditor conducting a NIST CSF 2.0 compliance interview.

Your task: formulate ONE clear, conversational question to ask the interviewee about the following NIST subcategory.

SUBCATEGORY DETAILS:
- ID: {subcategory_id}
- Category: {category}
- Description: {description}
- Implementation Examples: {examples}
{sp800_section}
INSTRUCTIONS:
1. Ask about their CURRENT state regarding this subcategory
2. Be specific but not overwhelming — ask one focused question
3. Briefly explain what this subcategory is about so the interviewee understands the context
4. The question MUST explicitly ask the interviewee to provide information about ALL of the following areas for the COMPANY/ENTITY being profiled:
   a) Whether this area is applicable/relevant to their organization
   b) The PRIORITY they assign to this area (High, Medium, or Low)
   c) Their current implementation status
   d) Any formal policies, processes, or procedures in place
   e) Internal or informal practices being followed
   f) Who is responsible (roles and responsibilities)
   g) Any standards, frameworks, or informative references they follow for this area (e.g. ISO 27001, NIST SP 800-53, sector-specific regulations, or any other relevant standard for their industry/context)
   h) Any documentary evidence or artifacts they can provide
5. Keep the tone professional but approachable
6. Write ONLY the question, no preamble or extra text

Question:"""

        messages = [
            {
                "role": "system",
                "content": "You are a cybersecurity compliance auditor. Generate clear, professional interview questions.",
            },
            {"role": "user", "content": prompt},
        ]

        trace = {
            "prompt": prompt,
            "messages": messages,
            "model": self.model_name,
        }

        self._print_verbose("\n🤖 [Question Generation] Sending prompt to LLM...")
        self._print_verbose(f"   Prompt length: {len(prompt)} chars")

        try:
            response = self.client.chat.completions.create(
                model=self.model_name,
                messages=messages,
                temperature=0.7,
                max_tokens=900,
            )

            content = response.choices[0].message.content or ""
            reasoning = getattr(response.choices[0].message, "reasoning_content", None)

            trace["response"] = {
                "content": content,
                "reasoning_content": reasoning,
                "tokens": {
                    "input": response.usage.prompt_tokens,
                    "output": response.usage.completion_tokens,
                },
            }

            self._print_verbose(f"   ✅ Response: {len(content)} chars")
            self._print_verbose(f"   Tokens — in: {response.usage.prompt_tokens}, out: {response.usage.completion_tokens}")
            if reasoning:
                self._print_verbose(f"   🧠 Reasoning: {reasoning[:200]}...")

            return content.strip(), trace

        except Exception as e:
            self._print_verbose(f"   ❌ LLM error: {e}")
            trace["error"] = str(e)
            # Fallback: use a generic question
            fallback = (
                f"Please describe your organization's current practices regarding "
                f"'{description}' ({subcategory_id}). "
                f"Include: priority (High/Medium/Low), policies, practices, "
                f"responsibilities, informative references, and evidence."
            )
            return fallback, trace

    def _build_target_question(
        self, row: Dict, catalog_ctx: Optional[Dict], current_state: Dict[str, str]
    ) -> tuple:
        """
        Uses the LLM to formulate a clear, professional interview question
        for the TARGET phase, taking into consideration the CURRENT state
        just collected.

        Returns:
            (question_text, trace_dict)
        """
        subcategory_id = row.get("Subcategory_ID", "")
        category = row.get("Category", "")
        description = row.get("Subcategory_Description", "")

        # Format current state cleanly
        current_state_text = "\n".join(
            f"  - {k.replace('Current_', '')}: {v}"
            for k, v in current_state.items()
            if v and v != "Not specified" and k.startswith("Current_")
        )
        if not current_state_text:
            current_state_text = "  (No specific current state information was provided by the user)"

        prompt = f"""You are a professional cybersecurity auditor conducting a NIST CSF 2.0 compliance interview.
We are now in the TARGET phase of the interview for the following subcategory.

SUBCATEGORY DETAILS:
- ID: {subcategory_id}
- Category: {category}
- Description: {description}

CURRENT STATE (just collected from the user):
{current_state_text}

INSTRUCTIONS:
1. Formulate ONE clear, conversational question asking about the organization's TARGET state (goals, improvements, desired future state) for this subcategory.
2. Acknowledge their current state briefly to make the question contextual (e.g., "Given that you currently do X...").
3. The question MUST explicitly ask the user for:
   a) Target Priority (High, Medium, Low)
   b) Target CSF Tier (Tier 1 to 4)
   c) Target Policies, Processes, or Procedures they want to implement
   d) Target Internal Practices they aim for
   e) Target Roles and Responsibilities (any changes needed?)
   f) Target Selected Informative References (new frameworks/standards to adopt?)
4. Keep the tone professional but approachable.
5. Write ONLY the question, no preamble or extra text.

Question:"""

        messages = [
            {
                "role": "system",
                "content": "You are a cybersecurity compliance auditor. Generate clear, professional interview questions.",
            },
            {"role": "user", "content": prompt},
        ]

        trace = {
            "prompt": prompt,
            "messages": messages,
            "model": self.model_name,
        }

        self._print_verbose("\n🤖 [TARGET Question Generation] Sending prompt to LLM...")
        
        try:
            response = self.client.chat.completions.create(
                model=self.model_name,
                messages=messages,
                temperature=0.7,
                max_tokens=900,
            )

            content = response.choices[0].message.content or ""
            reasoning = getattr(response.choices[0].message, "reasoning_content", None)

            trace["response"] = {
                "content": content,
                "reasoning_content": reasoning,
            }

            self._print_verbose(f"   ✅ Response: {len(content)} chars")
            return content.strip(), trace

        except Exception as e:
            self._print_verbose(f"   ❌ LLM error: {e}")
            trace["error"] = str(e)
            fallback = (
                f"Based on your current practices, what is your target state regarding "
                f"'{description}' ({subcategory_id})? "
                f"Please specify your Target Priority, Target Tier, and any planned policies, practices, or frameworks."
            )
            return fallback, trace

    # ═══════════════════════════════════════════════════════════════
    #  INTERNAL — Response extraction (with retry)
    # ═══════════════════════════════════════════════════════════════

    def _extract_response(
        self, row: Dict, user_answer: str,
        catalog_ctx: Optional[Dict] = None,
        previous_extracted: Optional[Dict] = None,
        feedback: Optional[str] = None
    ) -> tuple:
        """
        Uses the LLM to extract structured profile data from the user's
        free-text answer. Retries up to EXTRACTION_RETRIES times.
        If previous_extracted and feedback are provided, it refines the JSON.

        Returns:
            (extracted_dict, trace_dict)
        """
        subcategory_id = row.get("Subcategory_ID", "")
        category = row.get("Category", "")
        description = row.get("Subcategory_Description", "")

        # Build SP800-53 context for extraction guidance
        sp800_hint = ""
        if catalog_ctx and catalog_ctx.get("sp800_controls"):
            ctrl_list = ", ".join(
                f"{c['control_id']} ({c['control_name']})"
                for c in catalog_ctx["sp800_controls"]
            )
            sp800_hint = (
                f"\nNOTE: The related NIST SP 800-53 controls for this subcategory are: {ctrl_list}.\n"
                f"If the user mentions any of these or other standards/frameworks, include them in "
                f"Current_Selected_Informative_References.\n"
            )

        if previous_extracted and feedback:
            prompt = f"""Update the structured data based on the user's feedback.

SUBCATEGORY: {subcategory_id} ({description})
ORIGINAL ANSWER: "{user_answer}"
{sp800_hint}

PREVIOUSLY EXTRACTED DATA:
{json.dumps(previous_extracted, indent=2)}

USER FEEDBACK / CORRECTIONS:
"{feedback}"

Return ONLY this JSON updated with the user's feedback (no markdown, no explanation):

{{
    "Included_in_Profile": "Yes or No",
    "Rationale": "Brief justification",
    "Current_Priority": "High/Medium/Low, or 'Not specified'",
    "Current_Status": "Implementation status",
    "Current_Policies_Processes_Procedures": "Formal policies/procedures",
    "Current_Internal_Practices": "Informal practices and tools used",
    "Current_Roles_and_Responsibilities": "Who is responsible",
    "Current_Selected_Informative_References": "Only standards/frameworks the user explicitly named",
    "Current_Artifacts_and_Evidence": "Documentation and evidence"
}}"""
        else:
            prompt = f"""Extract structured data from this interview answer about NIST CSF 2.0 subcategory {subcategory_id} ({description}).
{sp800_hint}
ANSWER: "{user_answer}"

Return ONLY this JSON (no markdown, no explanation):

{{
    "Included_in_Profile": "Yes or No",
    "Rationale": "Brief justification",
    "Current_Priority": "High/Medium/Low, or 'Not specified' if user didn't say",
    "Current_Status": "Implementation status",
    "Current_Policies_Processes_Procedures": "Formal policies/procedures",
    "Current_Internal_Practices": "Informal practices and tools used",
    "Current_Roles_and_Responsibilities": "Who is responsible",
    "Current_Selected_Informative_References": "Only standards/frameworks/regulations the user explicitly named (e.g. ISO 27001, NIST SP 800-53). Put tools in Internal_Practices instead.",
    "Current_Artifacts_and_Evidence": "Documentation and evidence"
}}

Rules: Use 'Not specified' for missing fields. Only extract what the user actually said."""

        messages = [
            {
                "role": "system",
                "content": (
                    "You are a data extraction assistant. "
                    "Think briefly, then output ONLY valid JSON. "
                    "Do not over-analyze or repeat yourself."
                ),
            },
            {"role": "user", "content": prompt},
        ]

        trace = {
            "prompt": prompt,
            "messages": messages,
            "model": self.model_name,
            "user_answer": user_answer,
            "attempts": [],
        }

        self._print_verbose("\n🔍 [Response Extraction] Sending to LLM...")
        self._print_verbose(f"   User answer: {user_answer[:100]}...")

        # Retry loop: reasoning models may consume all tokens on reasoning
        for attempt in range(1, self.EXTRACTION_RETRIES + 1):
            attempt_trace = {"attempt": attempt}

            try:
                response = self.client.chat.completions.create(
                    model=self.model_name,
                    messages=messages,
                    temperature=0.0,
                    max_tokens=self.EXTRACTION_MAX_TOKENS,
                )

                raw_content = response.choices[0].message.content or ""
                reasoning = getattr(response.choices[0].message, "reasoning_content", None)
                content = raw_content.strip()

                attempt_trace["response"] = {
                    "content": raw_content,
                    "reasoning_content": reasoning,
                    "tokens": {
                        "input": response.usage.prompt_tokens,
                        "output": response.usage.completion_tokens,
                    },
                }

                self._print_verbose(
                    f"   [Attempt {attempt}/{self.EXTRACTION_RETRIES}] "
                    f"Raw response: {len(content)} chars, "
                    f"Tokens — in: {response.usage.prompt_tokens}, out: {response.usage.completion_tokens}"
                )
                if reasoning:
                    self._print_verbose(f"   🧠 Reasoning: {reasoning[:200]}...")

                # If content is empty, reasoning consumed all tokens → retry
                if not content:
                    self._print_verbose(
                        f"   ⚠️  Empty content (reasoning consumed all tokens). "
                        f"{'Retrying...' if attempt < self.EXTRACTION_RETRIES else 'Using fallback.'}"
                    )
                    attempt_trace["error"] = "Empty content — reasoning consumed all tokens"
                    trace["attempts"].append(attempt_trace)
                    continue

                self._print_verbose(f"   Raw content:\n{content}")

                # Clean markdown code blocks if present
                if "```json" in content:
                    content = content.split("```json")[1].split("```")[0].strip()
                elif "```" in content:
                    content = content.split("```")[1].split("```")[0].strip()

                extracted = json.loads(content)
                attempt_trace["parsed"] = extracted
                trace["attempts"].append(attempt_trace)

                # Validate: only keep known columns
                validated = {}
                for col in self.CURRENT_PROFILE_COLUMNS:
                    validated[col] = extracted.get(col, "Not specified")

                trace["success"] = True
                return validated, trace

            except json.JSONDecodeError as e:
                self._print_verbose(f"   ❌ JSON parse error: {e}")
                attempt_trace["error"] = f"JSONDecodeError: {e}"
                trace["attempts"].append(attempt_trace)
                break

            except Exception as e:
                self._print_verbose(f"   ❌ LLM error: {e}")
                attempt_trace["error"] = str(e)
                trace["attempts"].append(attempt_trace)
                break

        # All attempts exhausted → fallback
        self._print_verbose("   ⚠️  Using fallback extraction (raw answer saved).")
        trace["success"] = False
        return self._fallback_extraction(user_answer), trace

    def _fallback_extraction(self, user_answer: str) -> Dict[str, str]:
        """
        Fallback when LLM extraction fails: store the raw answer
        in Current_Status so nothing is lost.
        """
        return {
            "Included_in_Profile": "Yes",
            "Rationale": "Auto-included (extraction fallback)",
            "Current_Priority": "Not specified",
            "Current_Status": user_answer,
            "Current_Policies_Processes_Procedures": "Not specified",
            "Current_Internal_Practices": "Not specified",
            "Current_Roles_and_Responsibilities": "Not specified",
            "Current_Selected_Informative_References": "Not specified",
            "Current_Artifacts_and_Evidence": "Not specified",
        }

    def _extract_target_response(
        self, row: Dict, user_answer: str,
        catalog_ctx: Optional[Dict] = None,
        previous_extracted: Optional[Dict] = None,
        feedback: Optional[str] = None
    ) -> tuple:
        """
        Extract structured profile data from the user's free-text answer
        for the TARGET phase (populating Target_* columns).
        """
        subcategory_id = row.get("Subcategory_ID", "")
        description = row.get("Subcategory_Description", "")

        # Build SP800-53 context for extraction guidance (inspiration)
        sp800_hint = ""
        if catalog_ctx and catalog_ctx.get("sp800_controls"):
            ctrl_list = ", ".join(
                f"{c['control_id']} ({c['control_name']})"
                for c in catalog_ctx["sp800_controls"]
            )
            sp800_hint = (
                f"\nNOTE: The related NIST SP 800-53 controls for this subcategory are: {ctrl_list}.\n"
                f"You may use these for inspiration if the user's target answer implies them, "
                f"but stick primarily to what the user actually said. If the user explicitly mentions adopting these, "
                f"include them in Target_Selected_Informative_References.\n"
            )

        if previous_extracted and feedback:
            prompt = f"""Update the structured data about the TARGET STATE based on the user's feedback.

SUBCATEGORY: {subcategory_id} ({description})
ORIGINAL ANSWER: "{user_answer}"
{sp800_hint}

PREVIOUSLY EXTRACTED TARGET DATA:
{json.dumps(previous_extracted, indent=2)}

USER FEEDBACK / CORRECTIONS:
"{feedback}"

Return ONLY this JSON updated with the user's feedback (no markdown, no explanation):

{{
    "Target_Priority": "High/Medium/Low, or 'Not specified'",
    "Target_CSF_Tier": "Tier 1/2/3/4, or 'Not specified'",
    "Target_Policies_Processes_Procedures": "Target/planned policies",
    "Target_Internal_Practices": "Target/planned practices",
    "Target_Roles_and_Responsibilities": "Target responsibilities",
    "Target_Selected_Informative_References": "Target frameworks/standards to adopt"
}}"""
        else:
            prompt = f"""Extract structured data from this interview answer about the **TARGET STATE** (future goals) for NIST CSF 2.0 subcategory {subcategory_id} ({description}).
{sp800_hint}
ANSWER: "{user_answer}"

Return ONLY this JSON (no markdown, no explanation):

{{
    "Target_Priority": "High/Medium/Low, or 'Not specified'",
    "Target_CSF_Tier": "Tier 1/2/3/4, or 'Not specified'",
    "Target_Policies_Processes_Procedures": "Target/planned policies",
    "Target_Internal_Practices": "Target/planned practices",
    "Target_Roles_and_Responsibilities": "Target responsibilities",
    "Target_Selected_Informative_References": "Target frameworks/standards to adopt"
}}

Rules: Use 'Not specified' for missing fields. Only extract what the user actually said about their TARGET or FUTURE state."""

        messages = [
            {
                "role": "system",
                "content": (
                    "You are a data extraction assistant. "
                    "Think briefly, then output ONLY valid JSON. "
                    "Do not over-analyze or repeat yourself."
                ),
            },
            {"role": "user", "content": prompt},
        ]

        trace = {
            "prompt": prompt,
            "messages": messages,
            "model": self.model_name,
            "user_answer": user_answer,
            "attempts": [],
        }

        self._print_verbose("\n🔍 [TARGET Response Extraction] Sending to LLM...")

        for attempt in range(1, self.EXTRACTION_RETRIES + 1):
            attempt_trace = {"attempt": attempt}
            try:
                response = self.client.chat.completions.create(
                    model=self.model_name,
                    messages=messages,
                    temperature=0.0,
                    max_tokens=self.EXTRACTION_MAX_TOKENS,
                )

                raw_content = response.choices[0].message.content or ""
                reasoning = getattr(response.choices[0].message, "reasoning_content", None)
                content = raw_content.strip()
                
                if not content:
                    attempt_trace["error"] = "Empty content — reasoning consumed all tokens"
                    trace["attempts"].append(attempt_trace)
                    continue

                if "```json" in content:
                    content = content.split("```json")[1].split("```")[0].strip()
                elif "```" in content:
                    content = content.split("```")[1].split("```")[0].strip()

                extracted = json.loads(content)
                attempt_trace["parsed"] = extracted
                trace["attempts"].append(attempt_trace)

                validated = {}
                for col in self.TARGET_PROFILE_COLUMNS:
                    validated[col] = extracted.get(col, "Not specified")

                trace["success"] = True
                return validated, trace

            except Exception as e:
                attempt_trace["error"] = str(e)
                trace["attempts"].append(attempt_trace)
                if isinstance(e, json.JSONDecodeError):
                    break

        self._print_verbose("   ⚠️  Using fallback TARGET extraction (raw answer saved).")
        trace["success"] = False
        
        fallback_data = {
            "Target_Priority": "Not specified",
            "Target_CSF_Tier": "Not specified",
            "Target_Policies_Processes_Procedures": user_answer,
            "Target_Internal_Practices": "Not specified",
            "Target_Roles_and_Responsibilities": "Not specified",
            "Target_Selected_Informative_References": "Not specified",
        }
        return fallback_data, trace

    # ═══════════════════════════════════════════════════════════════
    #  INTERNAL — Progress & logging
    # ═══════════════════════════════════════════════════════════════

    def _show_progress(self):
        """Display current completion progress."""
        summary = self.manager.get_progress_summary()

        # Also compute in-scope progress
        if self.SCOPE_SUBCATEGORIES is not None:
            in_scope = self.manager.df[
                self.manager.df["Subcategory_ID"].isin(self.SCOPE_SUBCATEGORIES)
            ]
            # Count unique subcategory IDs (not rows)
            scope_done = in_scope[in_scope["Completion_Status"] == "DONE"][
                "Subcategory_ID"
            ].nunique()
            scope_total = in_scope["Subcategory_ID"].nunique()
            scope_pct = (scope_done / scope_total * 100) if scope_total > 0 else 0

            self._print_always(
                f"\n📊 Progress (in-scope): {scope_done}/{scope_total} "
                f"subcategories ({scope_pct:.0f}%)"
            )
        else:
            self._print_always(
                f"\n📊 Progress: {summary['completed']}/{summary['total_items']} "
                f"({summary['percentage']:.1f}%)"
            )

    def _save_run_log(self):
        """Save the full run log to a JSON file."""
        log_path = os.path.join(self.log_dir, "interview_log.json")
        log_data = {
            "run_timestamp": datetime.now().isoformat(),
            "model": self.model_name,
            "profile_path": self.manager.state_path,
            "scope": sorted(self.SCOPE_SUBCATEGORIES) if self.SCOPE_SUBCATEGORIES else "ALL",
            "turns": self._turn_logs,
        }
        with open(log_path, "w", encoding="utf-8") as f:
            json.dump(log_data, f, indent=2, ensure_ascii=False)

        self._print_verbose(f"   💾 Log saved: {log_path}")

    # ═══════════════════════════════════════════════════════════════
    #  INTERNAL — Output helpers
    # ═══════════════════════════════════════════════════════════════

    def _print_always(self, message: str):
        """Print regardless of verbose setting (user-facing messages)."""
        print(message)

    def _print_verbose(self, message: str):
        """Print only if verbose mode is on (internal traces)."""
        if self.verbose:
            print(message)


# ═══════════════════════════════════════════════════════════════════
#  CLI entry point
# ═══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="NIST CSF 2.0 Interview Engine"
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        default=False,
        help="Enable verbose mode (show all internal LLM traces)",
    )
    parser.add_argument(
        "--pandas-model",
        type=str,
        default="gpt-oss",
        help="LLM model name for Pandas reasoning (default: gpt-oss)",
    )
    args = parser.parse_args()

    # ── API Provider Selection Prompt ──
    print("\n" + "="*60)
    print("🌐 Select API Provider")
    print("="*60)
    print("  [1] UniBS Cluster (Free, Local)")
    print("  [2] OpenRouter (External via API Key)")
    print("  [3] GitHub Models (External via Azure API)")
    print("="*60)
    
    provider_choice = input("Select provider [1-3] (default: 1): ").strip()
    
    if provider_choice == "2":
        base_url = "https://openrouter.ai/api/v1"
        api_key = os.environ.get("OPENROUTER_API_KEY")
        models = [
            "openai/gpt-oss-20b:free",
            "google/gemini-2.5-flash",
            "meta-llama/llama-3.3-70b-instruct",
            "anthropic/claude-3.5-haiku",
            "mistralai/mistral-small-24b-instruct-2501",
            "deepseek/deepseek-chat"
        ]
    elif provider_choice == "3":
        base_url = "https://models.inference.ai.azure.com"
        api_key = os.environ.get("GITHUB_TOKEN")
        models = [
            "Phi-4",
            "gpt-4o",
            "gpt-4o-mini",
            "Meta-Llama-3-70B-Instruct"
        ]
    else:
        base_url = "https://gpustack.ing.unibs.it/v1"
        api_key = None
        models = ["qwen3", "phi4", "phi4-mini", "llama3.2", "gpt-oss", "granite3.3", "gemma3"]

    # ── Model Selection Prompt ──
    print("\n" + "="*60)
    print("🤖 Select Interview Model")
    print("="*60)
    for i, m in enumerate(models, 1):
        print(f"  [{i}] {m}")
    if provider_choice in ["2", "3"]:
        print(f"  [X] Type a custom model ID")
    print("="*60)
    
    choice = input(f"Select model [1-{len(models)}] or type ID (default: 1): ").strip()
    try:
        idx = int(choice) - 1
        if 0 <= idx < len(models):
            interview_model = models[idx]
        else:
            interview_model = models[0]
    except ValueError:
        if choice and provider_choice in ["2", "3"]:
            interview_model = choice
        else:
            interview_model = models[0]
            

    if provider_choice in ["2", "3"]:
        pandas_model = interview_model
    else:
        pandas_model = args.pandas_model
        
    print(f"\n✅ Using '{interview_model}' for interview and '{pandas_model}' for Pandas Agent.\n")

    engine = InterviewEngine(
        verbose=args.verbose,
        base_url=base_url,
        api_key=api_key,
        interview_model_name=interview_model,
        pandas_agent_model_name=pandas_model,
    )
    engine.start()
